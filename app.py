from flask import Flask, render_template, request, redirect, Response, session, abort
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, date, timedelta
from functools import wraps
from invoice_generator import generate_invoice
import sqlite3
import os
import csv
import io
import json
import secrets

app = Flask(__name__, template_folder="template")

_secret = os.environ.get("SECRET_KEY", "")
if not _secret:
    # Stable fallback — derive from DB path so same machine always gets same key
    # This prevents session invalidation on worker restarts (PythonAnywhere free tier)
    import hashlib
    _base  = os.path.abspath(__file__)
    _secret = hashlib.sha256((_base + "ratnakar_jewellery_stable_key_v7").encode()).hexdigest()
app.secret_key = _secret

SESSION_TIMEOUT_MINUTES = 120

# Secure session cookie settings (HTTPS enforcement)
app.config['SESSION_COOKIE_SECURE']   = True
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
MAX_ATTEMPTS            = 5
LOCKOUT_MINUTES         = 15

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
DB_NAME    = os.path.join(BASE_DIR, "jewellery.db")
BACKUP_DIR = os.path.join(BASE_DIR, "backups")



# ---------- NO CACHE ----------
@app.after_request
def no_cache(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
    response.headers["Pragma"]        = "no-cache"
    response.headers["Expires"]       = "0"
    return response


# ---------- DATABASE SETUP ----------
def init_db():
    conn = sqlite3.connect(DB_NAME)
    cur  = conn.cursor()
    cur.execute("""CREATE TABLE IF NOT EXISTS stock (
        ID TEXT PRIMARY KEY, ITEM TEXT NOT NULL, MATERIAL TEXT NOT NULL,
        CATEGORY TEXT NOT NULL, G_WEIGHT REAL NOT NULL DEFAULT 0, L_WEIGHT REAL NOT NULL DEFAULT 0, N_WEIGHT REAL NOT NULL DEFAULT 0, PURITY REAL,
        PURCHASE_DATE TEXT NOT NULL, NOTES TEXT NOT NULL DEFAULT '',
        CENT REAL DEFAULT NULL, MRP_PRICE REAL DEFAULT NULL)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS purchase (
        ID TEXT PRIMARY KEY, ITEM TEXT NOT NULL, MATERIAL TEXT NOT NULL,
        CATEGORY TEXT NOT NULL, G_WEIGHT REAL NOT NULL DEFAULT 0, L_WEIGHT REAL NOT NULL DEFAULT 0, N_WEIGHT REAL NOT NULL DEFAULT 0, PURITY REAL,
        SELLER TEXT NOT NULL, PHONE TEXT NOT NULL, PURCHASE_DATE TEXT NOT NULL,
        CENT REAL DEFAULT NULL, MRP_PRICE REAL DEFAULT NULL)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS sale (
        ID INTEGER PRIMARY KEY AUTOINCREMENT, ITEM TEXT NOT NULL, MATERIAL TEXT NOT NULL,
        CATEGORY TEXT NOT NULL, WEIGHT REAL NOT NULL, PURITY REAL,
        BUYER TEXT NOT NULL, PHONE TEXT NOT NULL, SALE_DATE TEXT NOT NULL,
        BUYER_ADDRESS TEXT NOT NULL DEFAULT '', MRP_PRICE REAL DEFAULT NULL)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS users (
        ID INTEGER PRIMARY KEY AUTOINCREMENT, USERNAME TEXT NOT NULL UNIQUE,
        PASSWORD TEXT NOT NULL, ROLE TEXT NOT NULL DEFAULT 'user')""")
    cur.execute("""CREATE TABLE IF NOT EXISTS invoice_data (
        SALE_ID INTEGER PRIMARY KEY, BUYER_NAME TEXT NOT NULL, BUYER_PHONE TEXT NOT NULL,
        BUYER_STATE TEXT NOT NULL, BUYER_GSTIN TEXT, PAYMENT_METHOD TEXT NOT NULL,
        SALE_DATE TEXT NOT NULL, ITEMS_JSON TEXT NOT NULL)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS payments (
        ID INTEGER PRIMARY KEY AUTOINCREMENT, SALE_ID INTEGER NOT NULL,
        BUYER_NAME TEXT NOT NULL, BUYER_PHONE TEXT NOT NULL, TOTAL_AMOUNT REAL NOT NULL,
        PAID_AMOUNT REAL NOT NULL, DUE_AMOUNT REAL NOT NULL, SALE_DATE TEXT NOT NULL,
        LAST_PAYMENT_DATE TEXT NOT NULL, STATUS TEXT NOT NULL DEFAULT 'Pending')""")
    cur.execute("""CREATE TABLE IF NOT EXISTS trade_dues (
        ID INTEGER PRIMARY KEY AUTOINCREMENT, PURCHASE_DATE TEXT NOT NULL,
        SELLER_NAME TEXT NOT NULL, SELLER_PHONE TEXT NOT NULL, MATERIAL TEXT NOT NULL,
        AGREED_WEIGHT REAL NOT NULL, GIVEN_WEIGHT REAL NOT NULL, DUE_WEIGHT REAL NOT NULL,
        LAST_UPDATE TEXT NOT NULL, STATUS TEXT NOT NULL DEFAULT 'Pending')""")
    cur.execute("""CREATE TABLE IF NOT EXISTS login_attempts (
        ID INTEGER PRIMARY KEY AUTOINCREMENT, USERNAME TEXT NOT NULL,
        ATTEMPT_TIME TEXT NOT NULL, SUCCESS INTEGER NOT NULL DEFAULT 0)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS audit_log (
        ID INTEGER PRIMARY KEY AUTOINCREMENT,
        USERNAME TEXT NOT NULL,
        ACTION TEXT NOT NULL,
        DETAIL TEXT,
        TIMESTAMP TEXT NOT NULL
    )""")
    # Add ROLE column if upgrading from old DB without it
    try:
        cur.execute("ALTER TABLE users ADD COLUMN ROLE TEXT NOT NULL DEFAULT 'user'")
    except Exception:
        pass
    # Add weight columns if upgrading from old DB
    for tbl in ("stock", "purchase"):
        for col, default in (("G_WEIGHT","0"),("L_WEIGHT","0"),("N_WEIGHT","0")):
            try:
                cur.execute(f"ALTER TABLE {tbl} ADD COLUMN {col} REAL NOT NULL DEFAULT {default}")
            except Exception:
                pass
    # Add ADDRESS column to sale if upgrading
    try:
        cur.execute("ALTER TABLE sale ADD COLUMN BUYER_ADDRESS TEXT NOT NULL DEFAULT ''")
    except Exception:
        pass
    try:
        cur.execute("ALTER TABLE stock ADD COLUMN NOTES TEXT NOT NULL DEFAULT ''")
    except Exception:
        pass
    # Add CENT column for diamond items
    for tbl in ("stock", "purchase"):
        try:
            cur.execute(f"ALTER TABLE {tbl} ADD COLUMN CENT REAL DEFAULT NULL")
        except Exception:
            pass
    # Add MRP_PRICE column for silver MRP sales
    for tbl in ("stock", "purchase"):
        try:
            cur.execute(f"ALTER TABLE {tbl} ADD COLUMN MRP_PRICE REAL DEFAULT NULL")
        except Exception:
            pass
    try:
        cur.execute("ALTER TABLE sale ADD COLUMN MRP_PRICE REAL DEFAULT NULL")
    except Exception:
        pass
    if cur.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0:
        cur.execute("INSERT INTO users (USERNAME,PASSWORD,ROLE) VALUES (?,?,?)",
            ('admin', generate_password_hash('admin123'), 'admin'))
    conn.commit()
    conn.close()


# ---------- DB HELPER ----------
def get_db():
    conn = sqlite3.connect(DB_NAME, timeout=30)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.row_factory = sqlite3.Row
    return conn


# ---------- AUDIT LOG ----------
def audit(action, detail=""):
    try:
        conn = get_db()
        conn.execute(
            "INSERT INTO audit_log (USERNAME,ACTION,DETAIL,TIMESTAMP) VALUES (?,?,?,?)",
            (session.get("username","system"), action, detail, datetime.now().isoformat())
        )
        conn.commit()
        conn.close()
    except Exception:
        pass  # Never let audit logging crash the app


# ---------- PHONE HELPER ----------
def clean_phone(raw):
    """Strip +91 / 91 prefix, spaces, dashes. Return 10-digit string or original."""
    p = raw.strip().replace(" ","").replace("-","")
    if p.startswith("+91"):
        p = p[3:]
    elif p.startswith("91") and len(p) == 12:
        p = p[2:]
    return p


# ---------- CSRF ----------
def generate_csrf():
    if "csrf_token" not in session:
        session["csrf_token"] = secrets.token_hex(32)
    return session["csrf_token"]

def validate_csrf():
    token = request.form.get("csrf_token", "")
    if not token or token != session.get("csrf_token", ""):
        abort(403)

@app.context_processor
def inject_globals():
    return dict(csrf_token=generate_csrf(), current_year=date.today().year,
                session_role=session.get("role", "user"))


# ---------- SESSION TIMEOUT ----------
@app.before_request
def check_session_timeout():
    if session.get("logged_in"):
        last = session.get("last_active")
        if last:
            elapsed = (datetime.now() - datetime.fromisoformat(last)).total_seconds() / 60
            if elapsed > SESSION_TIMEOUT_MINUTES:
                session.clear()
                return redirect("/login?reason=timeout")
        session["last_active"] = datetime.now().isoformat()


# ---------- AUTH DECORATORS ----------
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect("/login")
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect("/login")
        if session.get("role") != "admin":
            abort(403)
        return f(*args, **kwargs)
    return decorated


# ---------- LOCKOUT HELPERS ----------
def is_locked_out(username):
    conn   = get_db()
    cutoff = (datetime.now() - timedelta(minutes=LOCKOUT_MINUTES)).isoformat()
    fails  = conn.execute(
        "SELECT COUNT(*) FROM login_attempts WHERE USERNAME=? AND SUCCESS=0 AND ATTEMPT_TIME>?",
        (username, cutoff)).fetchone()[0]
    conn.close()
    return fails >= MAX_ATTEMPTS

def record_attempt(username, success):
    conn = get_db()
    conn.execute("INSERT INTO login_attempts (USERNAME,ATTEMPT_TIME,SUCCESS) VALUES (?,?,?)",
        (username, datetime.now().isoformat(), 1 if success else 0))
    conn.execute("DELETE FROM login_attempts WHERE ATTEMPT_TIME<?",
        ((datetime.now() - timedelta(days=1)).isoformat(),))
    conn.commit()
    conn.close()

def remaining_attempts(username):
    conn   = get_db()
    cutoff = (datetime.now() - timedelta(minutes=LOCKOUT_MINUTES)).isoformat()
    fails  = conn.execute(
        "SELECT COUNT(*) FROM login_attempts WHERE USERNAME=? AND SUCCESS=0 AND ATTEMPT_TIME>?",
        (username, cutoff)).fetchone()[0]
    conn.close()
    return max(0, MAX_ATTEMPTS - fails)


# ---------- LOGIN ----------
@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.args.get("reason") == "timeout":
        error = "Session expired due to inactivity. Please log in again."

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        if is_locked_out(username):
            error = f"Too many failed attempts. Try again in {LOCKOUT_MINUTES} minutes."
            return render_template("login.html", error=error)

        conn = get_db()
        user = conn.execute("SELECT * FROM users WHERE USERNAME=?", (username,)).fetchone()
        conn.close()

        if user and check_password_hash(user["PASSWORD"], password):
            record_attempt(username, True)
            session["logged_in"]   = True
            session["username"]    = username
            session["role"]        = user["ROLE"]
            session["last_active"] = datetime.now().isoformat()
            generate_csrf()
            return redirect("/")
        else:
            record_attempt(username, False)
            left = remaining_attempts(username)
            if left == 0:
                error = f"Too many failed attempts. Try again in {LOCKOUT_MINUTES} minutes."
            else:
                error = f"Invalid username or password. {left} attempt(s) remaining."

    return render_template("login.html", error=error)


# ---------- LOGOUT ----------
@app.route("/logout", methods=["POST"])
def logout():
    validate_csrf()
    session.clear()
    return redirect("/login")


# ---------- CHANGE PASSWORD ----------
@app.route("/change_password", methods=["GET", "POST"])
@login_required
def change_password():
    error = success = None
    if request.method == "POST":
        validate_csrf()
        current = request.form.get("current_password", "").strip()
        new_pw  = request.form.get("new_password", "").strip()
        confirm = request.form.get("confirm_password", "").strip()
        conn    = get_db()
        user    = conn.execute("SELECT * FROM users WHERE USERNAME=?", (session["username"],)).fetchone()
        if not user or not check_password_hash(user["PASSWORD"], current):
            error = "Current password is incorrect."
        elif len(new_pw) < 6:
            error = "New password must be at least 6 characters."
        elif new_pw != confirm:
            error = "New passwords do not match."
        else:
            conn.execute("UPDATE users SET PASSWORD=? WHERE USERNAME=?",
                (generate_password_hash(new_pw), session["username"]))
            conn.commit()
            success = "Password changed successfully."
        conn.close()
    return render_template("change_password.html", error=error, success=success)


# ---------- MANAGE USERS (admin only) ----------
@app.route("/manage_users")
@admin_required
def manage_users():
    conn  = get_db()
    users = conn.execute("SELECT ID,USERNAME,ROLE FROM users ORDER BY ID").fetchall()
    conn.close()
    return render_template("manage_users.html", users=users, error=None)

@app.route("/add_user", methods=["POST"])
@admin_required
def add_user():
    validate_csrf()
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "").strip()
    role     = request.form.get("role", "user").strip()
    error    = None

    if not username or not password:
        error = "Username and password are required."
    elif len(password) < 6:
        error = "Password must be at least 6 characters."
    elif role not in ("admin", "user"):
        error = "Invalid role."
    else:
        try:
            conn = get_db()
            conn.execute("INSERT INTO users (USERNAME,PASSWORD,ROLE) VALUES (?,?,?)",
                (username, generate_password_hash(password), role))
            conn.commit()
            conn.close()
            audit("ADD_USER", f"User '{username}' added with role '{role}'")
            return redirect("/manage_users")
        except sqlite3.IntegrityError:
            error = f"Username '{username}' already exists."

    conn  = get_db()
    users = conn.execute("SELECT ID,USERNAME,ROLE FROM users ORDER BY ID").fetchall()
    conn.close()
    return render_template("manage_users.html", users=users, error=error)

@app.route("/delete_user/<int:user_id>", methods=["POST"])
@admin_required
def delete_user(user_id):
    validate_csrf()
    conn = get_db()
    me   = conn.execute("SELECT ID FROM users WHERE USERNAME=?", (session["username"],)).fetchone()
    if not (me and me["ID"] == user_id):
        conn.execute("DELETE FROM users WHERE ID=?", (user_id,))
        conn.commit()
        audit("DELETE_USER", f"User ID {user_id} deleted")
    conn.close()
    return redirect("/manage_users")

@app.route("/reset_user_password/<int:user_id>", methods=["POST"])
@admin_required
def reset_user_password(user_id):
    validate_csrf()
    new_pw = request.form.get("new_password", "").strip()
    conn   = get_db()
    users  = conn.execute("SELECT ID,USERNAME,ROLE FROM users ORDER BY ID").fetchall()
    if len(new_pw) < 6:
        conn.close()
        return render_template("manage_users.html", users=users,
            error="Password must be at least 6 characters.")
    conn.execute("UPDATE users SET PASSWORD=? WHERE ID=?",
        (generate_password_hash(new_pw), user_id))
    conn.commit()
    audit("RESET_PASSWORD", f"Password reset for user ID {user_id}")
    conn.close()
    return redirect("/manage_users")


# ---------- BACKUP AS XLSX ----------
@app.route("/backup_db")
@admin_required
def backup_db():
    """
    DB FORMAT — Full Backup (Export)
    =================================
    Generates a multi-sheet .xlsx file. Each sheet includes:
      Row 1 : Instruction note (import type + required columns)
      Row 2 : Column headers
      Row 3+: Data rows

    Sheets & Columns:
    ─────────────────────────────────────────────────────────────────
    STOCK        : ID, ITEM, MATERIAL, CATEGORY, G_WEIGHT, L_WEIGHT,
                   N_WEIGHT, PURITY, CENT, MRP_PRICE, PURCHASE_DATE, NOTES
    PURCHASES    : ID, ITEM, MATERIAL, CATEGORY, G_WEIGHT, L_WEIGHT,
                   N_WEIGHT, PURITY, CENT, MRP_PRICE, SELLER, PHONE, PURCHASE_DATE
    SALES        : ID, ITEM, MATERIAL, CATEGORY, WEIGHT, PURITY,
                   MRP_PRICE, BUYER, PHONE, SALE_DATE  (ID ignored on re-import)
    DUE PAYMENTS : SALE_ID, BUYER_NAME, BUYER_PHONE, TOTAL_AMOUNT,
                   PAID_AMOUNT, DUE_AMOUNT, SALE_DATE,
                   LAST_PAYMENT_DATE, STATUS  (reference only)
    TRADE DUES   : SELLER_NAME, SELLER_PHONE, MATERIAL, AGREED_WEIGHT,
                   GIVEN_WEIGHT, DUE_WEIGHT, PURCHASE_DATE,
                   LAST_UPDATE, STATUS  (reference only)
    ─────────────────────────────────────────────────────────────────
    Import: Use /import_data  |  Supported formats: .csv, .xlsx
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    import io as _io

    conn = get_db()

    # Each tuple: (sheet_name, sql, columns, import_note)
    SHEETS = [
        (
            "Stock",
            "SELECT ID,ITEM,MATERIAL,CATEGORY,G_WEIGHT,L_WEIGHT,N_WEIGHT,PURITY,CENT,MRP_PRICE,PURCHASE_DATE,NOTES FROM stock ORDER BY PURCHASE_DATE DESC",
            ["ID","ITEM","MATERIAL","CATEGORY","G_WEIGHT","L_WEIGHT","N_WEIGHT","PURITY","CENT","MRP_PRICE","PURCHASE_DATE","NOTES"],
            "Import type: stock  |  Required: ID,ITEM,MATERIAL,CATEGORY,G_WEIGHT,L_WEIGHT,N_WEIGHT,PURITY,PURCHASE_DATE  Optional: CENT(diamond),MRP_PRICE(silver),NOTES"
        ),
        (
            "Purchases",
            "SELECT ID,ITEM,MATERIAL,CATEGORY,G_WEIGHT,L_WEIGHT,N_WEIGHT,PURITY,CENT,MRP_PRICE,SELLER,PHONE,PURCHASE_DATE FROM purchase ORDER BY PURCHASE_DATE DESC",
            ["ID","ITEM","MATERIAL","CATEGORY","G_WEIGHT","L_WEIGHT","N_WEIGHT","PURITY","CENT","MRP_PRICE","SELLER","PHONE","PURCHASE_DATE"],
            "Import type: purchase  |  Required: ID,ITEM,MATERIAL,CATEGORY,G_WEIGHT,L_WEIGHT,N_WEIGHT,PURITY,SELLER,PHONE,PURCHASE_DATE  Optional: CENT,MRP_PRICE"
        ),
        (
            "Sales",
            "SELECT ID,ITEM,MATERIAL,CATEGORY,WEIGHT,PURITY,MRP_PRICE,BUYER,PHONE,SALE_DATE FROM sale ORDER BY SALE_DATE DESC",
            ["ID","ITEM","MATERIAL","CATEGORY","WEIGHT","PURITY","MRP_PRICE","BUYER","PHONE","SALE_DATE"],
            "Import type: sale  |  Required: ITEM,MATERIAL,CATEGORY,WEIGHT,PURITY,BUYER,PHONE,SALE_DATE  Optional: MRP_PRICE  (ID ignored on import)"
        ),
        (
            "Due Payments",
            "SELECT SALE_ID,BUYER_NAME,BUYER_PHONE,TOTAL_AMOUNT,PAID_AMOUNT,DUE_AMOUNT,SALE_DATE,LAST_PAYMENT_DATE,STATUS FROM payments ORDER BY SALE_DATE DESC",
            ["SALE_ID","BUYER_NAME","BUYER_PHONE","TOTAL_AMOUNT","PAID_AMOUNT","DUE_AMOUNT","SALE_DATE","LAST_PAYMENT_DATE","STATUS"],
            "Reference only — not importable via Import Data page"
        ),
        (
            "Trade Dues",
            "SELECT SELLER_NAME,SELLER_PHONE,MATERIAL,AGREED_WEIGHT,GIVEN_WEIGHT,DUE_WEIGHT,PURCHASE_DATE,LAST_UPDATE,STATUS FROM trade_dues ORDER BY PURCHASE_DATE DESC",
            ["SELLER_NAME","SELLER_PHONE","MATERIAL","AGREED_WEIGHT","GIVEN_WEIGHT","DUE_WEIGHT","PURCHASE_DATE","LAST_UPDATE","STATUS"],
            "Reference only — not importable via Import Data page"
        ),
    ]

    HEADER_FILL  = PatternFill("solid", fgColor="8B0000")
    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
    NOTE_FILL    = PatternFill("solid", fgColor="FFF3CD")
    NOTE_FONT    = Font(italic=True, color="856404", size=9)
    CENTER       = Alignment(horizontal="center", vertical="center")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    for sheet_name, sql, columns, note in SHEETS:
        rows = conn.execute(sql).fetchall()
        ws   = wb.create_sheet(title=sheet_name)

        # Row 1: instruction note spanning all columns
        ws.append([note] + [""] * (len(columns) - 1))
        note_cell = ws.cell(row=1, column=1)
        note_cell.fill = NOTE_FILL
        note_cell.font = NOTE_FONT
        note_cell.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        ws.row_dimensions[1].height = 28

        # Row 2: column headers
        ws.append(columns)
        for col_idx, _ in enumerate(columns, start=1):
            cell = ws.cell(row=2, column=col_idx)
            cell.fill   = HEADER_FILL
            cell.font   = HEADER_FONT
            cell.alignment = CENTER
        ws.row_dimensions[2].height = 20

        # Data rows
        for row_idx, row in enumerate(rows, start=3):
            ws.append([row[col] for col in columns])
            # Alternate row shading for readability
            if row_idx % 2 == 0:
                light = PatternFill("solid", fgColor="F9F9F9")
                for col_idx in range(1, len(columns) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = light

        # Auto-fit column widths (cap at 40)
        for col_idx, col_name in enumerate(columns, start=1):
            max_len = len(col_name)
            for row in rows:
                val = str(row[col_name]) if row[col_name] is not None else ""
                max_len = max(max_len, len(val))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

        # Freeze panes below header row
        ws.freeze_panes = "A3"

    conn.close()

    # Write to memory buffer — no temp file on disk
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename  = f"RatnakarJewellery_Backup_{timestamp}.xlsx"
    audit("BACKUP", f"Full data backup downloaded as {filename}")

    return Response(buf.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"})


# ---------- DASHBOARD ----------
@app.route("/")
@login_required
def dashboard():
    conn          = get_db()
    cur           = conn.cursor()
    gold          = cur.execute("SELECT IFNULL(SUM(N_WEIGHT),0) FROM stock WHERE MATERIAL='GOLD'").fetchone()[0]
    gold_in_diamond = cur.execute("SELECT IFNULL(SUM(N_WEIGHT),0) FROM stock WHERE MATERIAL='DIAMOND'").fetchone()[0]
    gold          = round(gold + gold_in_diamond, 3)
    silver        = cur.execute("SELECT IFNULL(SUM(N_WEIGHT),0) FROM stock WHERE MATERIAL='SILVER'").fetchone()[0]
    diamond       = cur.execute("SELECT COUNT(*) FROM stock WHERE MATERIAL='DIAMOND'").fetchone()[0]
    pending_dues  = cur.execute("SELECT COUNT(*) FROM payments WHERE STATUS='Pending'").fetchone()[0]
    pending_trades= cur.execute("SELECT COUNT(*) FROM trade_dues WHERE STATUS='Pending'").fetchone()[0]
    total_stock   = cur.execute("SELECT COUNT(*) FROM stock").fetchone()[0]
    conn.close()
    # DB size monitoring
    try:
        db_size_bytes = os.path.getsize(DB_NAME)
        db_size_mb    = round(db_size_bytes / (1024 * 1024), 2)
        db_size_pct   = round((db_size_bytes / (512 * 1024 * 1024)) * 100, 1)
    except OSError:
        db_size_mb  = 0
        db_size_pct = 0
    return render_template("dashboard.html", gold_weight=gold, silver_weight=silver,
        diamond_count=diamond, pending_dues=pending_dues, pending_trades=pending_trades,
        total_stock=total_stock, db_size_mb=db_size_mb, db_size_pct=db_size_pct)


# ---------- STOCK SEARCH API (for sale page live lookup) ----------
@app.route("/api/stock_search")
@login_required
def api_stock_search():
    q    = request.args.get("q", "").strip()
    if not q:
        return {"results": []}
    like = f"%{q.lower()}%"
    conn = get_db()
    rows = conn.execute(
        "SELECT ID,ITEM,MATERIAL,CATEGORY,G_WEIGHT,L_WEIGHT,N_WEIGHT,PURITY,CENT,MRP_PRICE FROM stock"
        " WHERE LOWER(ID) LIKE ? OR LOWER(ITEM) LIKE ?"
        " OR LOWER(MATERIAL) LIKE ? OR LOWER(CATEGORY) LIKE ?"
        " ORDER BY ITEM LIMIT 20",
        (like, like, like, like)
    ).fetchall()
    conn.close()
    return {"results": [dict(r) for r in rows]}


# ---------- VIEW STOCK (paginated + search + filter) ----------
@app.route("/view_stock")
@login_required
def view_stock():
    PER_PAGE = 50
    try:
        page = max(1, int(request.args.get("page", 1)))
    except ValueError:
        page = 1
    q               = request.args.get("q", "").strip()
    material_filter = request.args.get("material", "").strip().upper()
    conn   = get_db()
    params = []
    where  = []
    if q:
        like = f"%{q.lower()}%"
        where.append("(LOWER(ID) LIKE ? OR LOWER(ITEM) LIKE ? OR LOWER(CATEGORY) LIKE ?)")
        params.extend([like, like, like])
    if material_filter:
        where.append("MATERIAL = ?")
        params.append(material_filter)
    where_clause = ("WHERE " + " AND ".join(where)) if where else ""
    total       = conn.execute(f"SELECT COUNT(*) FROM stock {where_clause}", params).fetchone()[0]
    total_pages = max(1, (total + PER_PAGE - 1) // PER_PAGE)
    page        = min(page, total_pages)
    stock = conn.execute(
        f"SELECT ID,ITEM,MATERIAL,CATEGORY,G_WEIGHT,L_WEIGHT,N_WEIGHT,PURITY,PURCHASE_DATE,NOTES,CENT,MRP_PRICE FROM stock {where_clause} ORDER BY PURCHASE_DATE DESC LIMIT ? OFFSET ?",
        params + [PER_PAGE, (page-1)*PER_PAGE]).fetchall()
    conn.close()
    return render_template("view_stock.html", stock=stock,
        page=page, total_pages=total_pages, total=total, per_page=PER_PAGE,
        q=q, material_filter=material_filter)


# ---------- ADD PURCHASE ----------
@app.route("/add_purchase")
@login_required
def add_purchase():
    return render_template("add_purchase.html", pcart=session.get("pcart", []),
        today=datetime.now().strftime("%Y-%m-%d"), error=None)

@app.route("/add_to_pcart", methods=["POST"])
@login_required
def add_to_pcart():
    validate_csrf()
    tag_id   = request.form.get("tag_id", "").strip()[:50]
    item     = request.form.get("item", "").strip()[:100]
    material = request.form.get("material", "").strip().upper()[:20]
    category = request.form.get("category", "").strip().upper()[:50]
    error    = None
    try:
        g_weight = round(float(request.form.get("g_weight", 0)), 3)
        l_weight = round(float(request.form.get("l_weight", 0)), 3)
        n_weight = round(g_weight - l_weight, 3)
        # Purity: optional (None) for SILVER, required for others
        purity_raw = request.form.get("purity", "").strip()
        if material == "SILVER" and purity_raw == "":
            purity = None
        else:
            purity = float(purity_raw) if purity_raw else 0.0
        # Cent: only for DIAMOND
        cent_raw = request.form.get("cent", "").strip()
        cent = float(cent_raw) if material == "DIAMOND" and cent_raw else None
        # MRP: only for SILVER
        mrp_raw = request.form.get("mrp_price", "").strip()
        mrp_price = float(mrp_raw) if material == "SILVER" and mrp_raw else None
    except ValueError:
        error = "Weight and Purity must be valid numbers."
    if not error and (not tag_id or not item or not material or not category):
        error = "All fields are required."
    if not error and material != "SILVER":
        # Non-silver: weight must be positive
        if g_weight <= 0:
            error = "G.Weight must be positive."
    if not error and material == "SILVER":
        # Silver with MRP: weight can be 0 (MRP items may not have weight)
        # Silver without MRP: still needs positive weight
        if mrp_price is None and g_weight <= 0:
            error = "G.Weight must be positive (or set MRP Price for fixed-price silver items)."
    if not error and material != "SILVER" and (purity is None or purity <= 0):
        error = "Purity must be positive for Gold and Diamond items."
    if not error and material != "SILVER" and n_weight <= 0:
        error = "N.Weight (G.Weight - L.Weight) must be greater than zero."
    if not error and material == "SILVER" and mrp_price is None and n_weight <= 0:
        error = "N.Weight must be greater than zero (or set MRP Price for fixed-price items)."
    if not error and g_weight > 5000:
        error = "G.Weight cannot exceed 5000g. Please check your entry."
    if not error:
        conn     = get_db()
        existing = conn.execute("SELECT ID FROM stock WHERE ID=?", (tag_id,)).fetchone()
        conn.close()
        if existing:
            error = f"Tag ID '{tag_id}' already exists in stock."
    if not error:
        pcart = session.get("pcart", [])
        if any(p["tag_id"] == tag_id for p in pcart):
            error = f"Tag ID '{tag_id}' already in cart."
    if error:
        return render_template("add_purchase.html", pcart=session.get("pcart", []),
            today=datetime.now().strftime("%Y-%m-%d"), error=error)
    pcart = session.get("pcart", [])
    pcart.append({"tag_id": tag_id, "item": item, "material": material,
                  "category": category, "g_weight": g_weight, "l_weight": l_weight,
                  "n_weight": n_weight, "purity": purity,
                  "cent": cent, "mrp_price": mrp_price})
    session["pcart"]  = pcart
    session.modified  = True
    return redirect("/add_purchase")

@app.route("/remove_from_pcart/<path:tag_id>")
@login_required
def remove_from_pcart(tag_id):
    session["pcart"]  = [i for i in session.get("pcart", []) if i["tag_id"] != tag_id]
    session.modified  = True
    return redirect("/add_purchase")

@app.route("/complete_purchase", methods=["POST"])
@login_required
def complete_purchase():
    validate_csrf()
    pcart = session.get("pcart", [])
    if not pcart:
        return redirect("/add_purchase")
    seller        = request.form.get("seller", "").strip()[:100]
    phone         = request.form.get("phone", "").strip()[:15]
    purchase_date = request.form.get("purchase_date", "") or datetime.now().strftime("%Y-%m-%d")
    if not seller:
        return render_template("add_purchase.html", pcart=pcart,
            today=datetime.now().strftime("%Y-%m-%d"), error="Seller name is required.")
    phone = clean_phone(phone)
    if not phone.isdigit() or len(phone) != 10:
        return render_template("add_purchase.html", pcart=pcart,
            today=datetime.now().strftime("%Y-%m-%d"), error="Phone must be exactly 10 digits (you can include +91).")
    payment_mode = request.form.get("payment_mode", "cash").strip()
    trade_items  = []
    if payment_mode in ("trade", "mixed"):
        for mat, aw, gw in zip(request.form.getlist("trade_material"),
                               request.form.getlist("trade_agreed_weight"),
                               request.form.getlist("trade_given_weight")):
            try:
                aw = round(float(aw), 3)
                gw = min(round(float(gw), 3), aw)
            except ValueError:
                continue
            if mat and aw > 0:
                trade_items.append({"material": mat.upper(), "agreed_weight": aw,
                    "given_weight": gw, "due_weight": round(aw - gw, 3)})
        if not trade_items:
            return render_template("add_purchase.html", pcart=pcart,
                today=datetime.now().strftime("%Y-%m-%d"),
                error="Add at least one trade item.")
    conn = get_db()
    cur  = conn.cursor()
    try:
        for i in pcart:
            if cur.execute("SELECT ID FROM stock WHERE ID=?", (i["tag_id"],)).fetchone():
                raise ValueError(f"Tag ID '{i['tag_id']}' already exists in stock.")
        for i in pcart:
            cur.execute("INSERT INTO stock (ID,ITEM,MATERIAL,CATEGORY,G_WEIGHT,L_WEIGHT,N_WEIGHT,PURITY,PURCHASE_DATE,NOTES,CENT,MRP_PRICE) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                (i["tag_id"], i["item"], i["material"], i["category"], i["g_weight"], i["l_weight"], i["n_weight"], i["purity"], purchase_date, "", i.get("cent"), i.get("mrp_price")))
            cur.execute("INSERT INTO purchase (ID,ITEM,MATERIAL,CATEGORY,G_WEIGHT,L_WEIGHT,N_WEIGHT,PURITY,SELLER,PHONE,PURCHASE_DATE,CENT,MRP_PRICE) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (i["tag_id"], i["item"], i["material"], i["category"], i["g_weight"], i["l_weight"], i["n_weight"], i["purity"], seller, phone, purchase_date, i.get("cent"), i.get("mrp_price")))
        for t in trade_items:
            cur.execute("""INSERT INTO trade_dues
                (PURCHASE_DATE,SELLER_NAME,SELLER_PHONE,MATERIAL,AGREED_WEIGHT,GIVEN_WEIGHT,DUE_WEIGHT,LAST_UPDATE,STATUS)
                VALUES (?,?,?,?,?,?,?,?,?)""",
                (purchase_date, seller, phone, t["material"], t["agreed_weight"],
                 t["given_weight"], t["due_weight"], purchase_date,
                 "Cleared" if t["due_weight"] <= 0 else "Pending"))
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        return render_template("add_purchase.html", pcart=pcart,
            today=datetime.now().strftime("%Y-%m-%d"), error=f"Purchase failed: {str(e)}")
    finally:
        try:
            conn.close()
        except Exception:
            pass
    session["pcart"]  = []
    session.modified  = True
    audit("PURCHASE", f"{len(pcart)} items from {seller}")
    return redirect("/view_stock")


# ---------- TRADE DUES ----------
@app.route("/trade_dues")
@login_required
def trade_dues():
    PER_PAGE = 50
    try:
        page = max(1, int(request.args.get("page", 1)))
    except ValueError:
        page = 1
    q             = request.args.get("q", "").strip()
    status_filter = request.args.get("status", "").strip()
    conn   = get_db()
    params = []
    where  = []
    if q:
        like = f"%{q.lower()}%"
        where.append("(LOWER(SELLER_NAME) LIKE ? OR SELLER_PHONE LIKE ?)")
        params.extend([like, like])
    if status_filter:
        where.append("STATUS = ?")
        params.append(status_filter)
    where_clause = ("WHERE " + " AND ".join(where)) if where else ""
    total       = conn.execute(f"SELECT COUNT(*) FROM trade_dues {where_clause}", params).fetchone()[0]
    total_pages = max(1, (total + PER_PAGE - 1) // PER_PAGE)
    page        = min(page, total_pages)
    dues = conn.execute(
        f"SELECT * FROM trade_dues {where_clause} ORDER BY STATUS ASC, PURCHASE_DATE DESC LIMIT ? OFFSET ?",
        params + [PER_PAGE, (page-1)*PER_PAGE]).fetchall()
    summary = conn.execute("SELECT COUNT(*), IFNULL(SUM(DUE_WEIGHT),0) FROM trade_dues WHERE STATUS='Pending'").fetchone()
    pending_count        = summary[0]
    total_pending_weight = round(summary[1], 3)
    conn.close()
    return render_template("trade_dues.html", dues=dues,
        page=page, total_pages=total_pages, total=total,
        q=q, status_filter=status_filter,
        pending_count=pending_count, total_pending_weight=total_pending_weight)

@app.route("/settle_trade/<int:due_id>", methods=["POST"])
@login_required
def settle_trade(due_id):
    validate_csrf()
    conn = get_db()
    row  = conn.execute("SELECT * FROM trade_dues WHERE ID=?", (due_id,)).fetchone()
    if not row:
        conn.close()
        return redirect("/trade_dues")
    try:
        weight = round(float(request.form.get("weight", 0)), 3)
    except ValueError:
        conn.close()
        return redirect("/trade_dues")
    if weight <= 0:
        conn.close()
        return redirect("/trade_dues")
    new_given = round(row["GIVEN_WEIGHT"] + weight, 3)
    new_due   = max(0, round(row["AGREED_WEIGHT"] - new_given, 3))
    if new_due == 0:
        new_given = row["AGREED_WEIGHT"]
    conn.execute("UPDATE trade_dues SET GIVEN_WEIGHT=?,DUE_WEIGHT=?,LAST_UPDATE=?,STATUS=? WHERE ID=?",
        (new_given, new_due, datetime.now().strftime("%Y-%m-%d"),
         "Cleared" if new_due <= 0 else "Pending", due_id))
    conn.commit()
    conn.close()
    return redirect("/trade_dues")


# ---------- RECORD SALE ----------
@app.route("/record_sale")
@login_required
def record_sale():
    cart_items = session.get("cart", [])
    total = round(sum((i["weight"] * i["rate_per_gram"]) + i["making_charges"] for i in cart_items), 2)
    return render_template("record_sale.html", cart=cart_items, total=total,
        today=datetime.now().strftime("%Y-%m-%d"))

@app.route("/cart")
@login_required
def cart():
    return record_sale()

@app.route("/add_to_cart", methods=["POST"])
@login_required
def add_to_cart():
    validate_csrf()
    stock_id = request.form.get("stock_id", "").strip()
    hsn      = request.form.get("hsn", "").strip()
    try:
        making_charges = float(request.form.get("making_charges", 0))
        flat_price     = float(request.form.get("flat_price", 0) or 0)
        rate_per_gram  = float(request.form.get("rate_per_gram", 0) or 0)
    except ValueError:
        return redirect("/cart")
    if not stock_id or not hsn:
        return redirect("/cart")
    conn = get_db()
    item = conn.execute("SELECT * FROM stock WHERE ID=?", (stock_id,)).fetchone()
    conn.close()
    if not item:
        return redirect("/cart")
    is_diamond = item["MATERIAL"] == "DIAMOND"
    is_mrp_silver = (item["MATERIAL"] == "SILVER" and
                     item["MRP_PRICE"] is not None and
                     float(item["MRP_PRICE"] or 0) > 0)

    if is_diamond:
        # Diamond: gold part priced per gram + stone flat price
        if rate_per_gram <= 0 and flat_price <= 0:
            return redirect("/cart")
        gold_part  = round(item["N_WEIGHT"] * rate_per_gram, 2) if rate_per_gram > 0 else 0
        item_total = round(gold_part + flat_price + making_charges, 2)
    elif is_mrp_silver:
        # MRP Silver: item_total = MRP price (inclusive, no rate/making needed)
        item_total    = round(float(item["MRP_PRICE"]), 2)
        rate_per_gram = 0
        making_charges = 0
        flat_price    = 0
    else:
        if rate_per_gram <= 0 or rate_per_gram > 1000000:
            return redirect("/cart")
        item_total = round((item["N_WEIGHT"] * rate_per_gram) + making_charges, 2)
        flat_price = 0
    cart = session.get("cart", [])
    if any(c["stock_id"] == stock_id for c in cart):
        return redirect("/cart")
    cent_val = item["CENT"] if "CENT" in item.keys() else None
    cart.append({
        "stock_id": stock_id, "item": item["ITEM"], "material": item["MATERIAL"],
        "category": item["CATEGORY"], "weight": item["N_WEIGHT"], "purity": item["PURITY"],
        "g_weight": item["G_WEIGHT"], "l_weight": item["L_WEIGHT"],
        "hsn": hsn, "rate_per_gram": rate_per_gram, "making_charges": making_charges,
        "flat_price": flat_price,
        "item_total": item_total,
        "is_mrp": is_mrp_silver,
        "cent": cent_val,
        "mrp_price": item["MRP_PRICE"] if "MRP_PRICE" in item.keys() else None
    })
    session["cart"]  = cart
    session.modified = True
    return redirect("/cart")

@app.route("/remove_from_cart/<path:stock_id>")
@login_required
def remove_from_cart(stock_id):
    session["cart"]  = [i for i in session.get("cart", []) if i["stock_id"] != stock_id]
    session.modified = True
    return redirect("/cart")

@app.route("/complete_sale", methods=["POST"])
@login_required
def complete_sale():
    validate_csrf()
    cart = session.get("cart", [])
    if not cart:
        return redirect("/cart")
    buyer          = request.form.get("buyer", "").strip()[:100]
    phone          = request.form.get("phone", "").strip()[:15]
    buyer_address  = request.form.get("buyer_address", "").strip()[:200]
    buyer_gstin    = request.form.get("buyer_gstin", "").strip()[:20]
    payment_method = request.form.get("payment_method", "Cash").strip()[:20]
    sale_date_raw  = request.form.get("sale_date", "").strip()
    phone = clean_phone(phone)
    if not buyer or not phone.isdigit() or len(phone) != 10:
        return redirect("/cart")
    if sale_date_raw:
        try:
            obj               = datetime.strptime(sale_date_raw, "%Y-%m-%d")
            sale_date_display = obj.strftime("%d-%m-%Y")
            sale_date_db      = obj.strftime("%Y-%m-%d")
        except ValueError:
            sale_date_display = datetime.now().strftime("%d-%m-%Y")
            sale_date_db      = datetime.now().strftime("%Y-%m-%d")
    else:
        sale_date_display = datetime.now().strftime("%d-%m-%Y")
        sale_date_db      = datetime.now().strftime("%Y-%m-%d")
    payment_type = request.form.get("payment_type", "full").strip()
    try:
        paid_now = float(request.form.get("paid_amount", 0))
    except ValueError:
        paid_now = 0.0
    conn = get_db()
    cur  = conn.cursor()
    try:
        sale_ids = []
        for i in cart:
            if not cur.execute("SELECT ID FROM stock WHERE ID=?", (i["stock_id"],)).fetchone():
                conn.rollback()
                conn.close()
                return f"Item '{i['stock_id']}' no longer in stock. Clear cart and retry.", 400
            cur.execute("""INSERT INTO sale (ITEM,MATERIAL,CATEGORY,WEIGHT,PURITY,BUYER,PHONE,SALE_DATE,BUYER_ADDRESS,MRP_PRICE)
                VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (i["item"], i["material"], i["category"], i["weight"], i["purity"],
                 buyer, phone, sale_date_db, buyer_address, i.get("mrp_price")))
            sale_ids.append(cur.lastrowid)
            cur.execute("DELETE FROM stock WHERE ID=?", (i["stock_id"],))
        cur.execute("""INSERT OR REPLACE INTO invoice_data
            (SALE_ID,BUYER_NAME,BUYER_PHONE,BUYER_STATE,BUYER_GSTIN,PAYMENT_METHOD,SALE_DATE,ITEMS_JSON)
            VALUES (?,?,?,?,?,?,?,?)""",
            (sale_ids[0], buyer, phone, buyer_address, buyer_gstin,
             payment_method, sale_date_display, json.dumps(cart)))
        taxable   = round(sum(i["item_total"] for i in cart), 2)
        total_amt = round(taxable * 1.03, 2)
        paid_amt  = total_amt if payment_type == "full" else round(min(paid_now, total_amt), 2)
        due_amt   = round(total_amt - paid_amt, 2)
        if due_amt > 0:
            cur.execute("""INSERT INTO payments
                (SALE_ID,BUYER_NAME,BUYER_PHONE,TOTAL_AMOUNT,PAID_AMOUNT,DUE_AMOUNT,SALE_DATE,LAST_PAYMENT_DATE,STATUS)
                VALUES (?,?,?,?,?,?,?,?,?)""",
                (sale_ids[0], buyer, phone, total_amt, paid_amt, due_amt,
                 sale_date_db, sale_date_db, 'Pending'))
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        return f"Sale failed: {str(e)}. No changes made.", 500
    conn.close()
    invoice_path = None
    try:
        # Normalise cart items so invoice generator never hits a missing key
        for itm in cart:
            itm.setdefault("flat_price", 0)
            itm.setdefault("rate_per_gram", 0)
            itm.setdefault("making_charges", 0)
            itm.setdefault("cent", None)
            itm.setdefault("mrp_price", None)
        invoice_path = generate_invoice(sale_id=sale_ids[0], buyer_name=buyer,
            buyer_phone=phone, buyer_address=buyer_address, buyer_gstin=buyer_gstin,
            payment_method=payment_method, sale_date=sale_date_display, items=cart)
        with open(invoice_path, "rb") as f:
            pdf_data = f.read()
    except Exception:
        pdf_data = None
    finally:
        if invoice_path and os.path.exists(invoice_path):
            try:
                os.remove(invoice_path)
            except OSError:
                pass
    session["cart"]  = []
    session.modified = True
    audit("SALE", f"Sale #{sale_ids[0]} to {buyer} — Rs.{round(sum(i['item_total'] for i in cart)*1.03,2)}")
    if pdf_data:
        return Response(pdf_data, mimetype="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=Invoice_{sale_ids[0]}_{buyer.replace(' ','_')}.pdf"})
    return redirect("/view_sales")


# ---------- VIEW PURCHASES (paginated) ----------
@app.route("/view_purchases")
@login_required
def view_purchases():
    PER_PAGE = 50
    try:
        page = max(1, int(request.args.get("page", 1)))
    except ValueError:
        page = 1
    conn        = get_db()
    total       = conn.execute("SELECT COUNT(*) FROM purchase").fetchone()[0]
    total_pages = max(1, (total + PER_PAGE - 1) // PER_PAGE)
    page        = min(page, total_pages)
    purchases   = conn.execute(
        "SELECT ID,ITEM,MATERIAL,CATEGORY,G_WEIGHT,L_WEIGHT,N_WEIGHT,PURITY,CENT,MRP_PRICE,SELLER,PHONE,PURCHASE_DATE FROM purchase ORDER BY PURCHASE_DATE DESC LIMIT ? OFFSET ?",
        (PER_PAGE, (page-1)*PER_PAGE)).fetchall()
    conn.close()
    return render_template("view_purchases.html", purchases=purchases,
        page=page, total_pages=total_pages, total=total, per_page=PER_PAGE)


# ---------- VIEW SALES (paginated + search) ----------
@app.route("/view_sales")
@login_required
def view_sales():
    PER_PAGE = 50
    try:
        page = max(1, int(request.args.get("page", 1)))
    except ValueError:
        page = 1
    q      = request.args.get("q", "").strip()
    conn   = get_db()
    params = []
    where  = []
    if q:
        like = f"%{q.lower()}%"
        where.append("(LOWER(s.ITEM) LIKE ? OR LOWER(s.BUYER) LIKE ? OR s.PHONE LIKE ? OR LOWER(s.MATERIAL) LIKE ?)")
        params.extend([like, like, like, like])
    where_clause = ("WHERE " + " AND ".join(where)) if where else ""
    total       = conn.execute(f"SELECT COUNT(*) FROM sale s {where_clause}", params).fetchone()[0]
    total_pages = max(1, (total + PER_PAGE - 1) // PER_PAGE)
    page        = min(page, total_pages)
    sales = conn.execute(f"""
        SELECT s.ID,s.ITEM,s.MATERIAL,s.CATEGORY,s.WEIGHT,s.PURITY,s.BUYER,s.PHONE,s.SALE_DATE,
               CASE WHEN i.SALE_ID IS NOT NULL THEN 1 ELSE 0 END as HAS_INVOICE
        FROM sale s LEFT JOIN invoice_data i ON s.ID=i.SALE_ID
        {where_clause}
        ORDER BY s.SALE_DATE DESC LIMIT ? OFFSET ?""",
        params + [PER_PAGE, (page-1)*PER_PAGE]).fetchall()
    conn.close()
    return render_template("view_sales.html", sales=sales,
        page=page, total_pages=total_pages, total=total, per_page=PER_PAGE, q=q)


# ---------- REPORTS (with totals + date range + material breakdown) ----------
@app.route("/report", methods=["GET", "POST"])
@login_required
def report():
    conn        = get_db()
    today       = date.today().strftime("%Y-%m-%d")
    report_type = request.form.get("type", "daily")
    date_from   = request.form.get("date_from", "").strip()
    date_to     = request.form.get("date_to", "").strip()

    if report_type == "custom" and date_from and date_to:
        purchase_q = ("SELECT ID,ITEM,MATERIAL,CATEGORY,G_WEIGHT,L_WEIGHT,N_WEIGHT,SELLER,PURCHASE_DATE FROM purchase WHERE PURCHASE_DATE BETWEEN ? AND ? ORDER BY PURCHASE_DATE DESC", (date_from, date_to))
        sale_q     = ("SELECT ITEM,MATERIAL,CATEGORY,WEIGHT,BUYER,SALE_DATE FROM sale WHERE SALE_DATE BETWEEN ? AND ? ORDER BY SALE_DATE DESC", (date_from, date_to))
        inv_q      = ("SELECT i.ITEMS_JSON FROM invoice_data i JOIN sale s ON s.ID=i.SALE_ID WHERE s.SALE_DATE BETWEEN ? AND ?", (date_from, date_to))
    else:
        QUERIES = {
            "daily":   {
                "p": ("SELECT ID,ITEM,MATERIAL,CATEGORY,G_WEIGHT,L_WEIGHT,N_WEIGHT,SELLER,PURCHASE_DATE FROM purchase WHERE PURCHASE_DATE=? ORDER BY PURCHASE_DATE DESC", (today,)),
                "s": ("SELECT ITEM,MATERIAL,CATEGORY,WEIGHT,BUYER,SALE_DATE FROM sale WHERE SALE_DATE=? ORDER BY SALE_DATE DESC", (today,)),
                "i": ("SELECT i.ITEMS_JSON FROM invoice_data i JOIN sale s ON s.ID=i.SALE_ID WHERE s.SALE_DATE=?", (today,)),
            },
            "monthly": {
                "p": ("SELECT ID,ITEM,MATERIAL,CATEGORY,G_WEIGHT,L_WEIGHT,N_WEIGHT,SELLER,PURCHASE_DATE FROM purchase WHERE strftime('%Y-%m',PURCHASE_DATE)=strftime('%Y-%m','now') ORDER BY PURCHASE_DATE DESC", ()),
                "s": ("SELECT ITEM,MATERIAL,CATEGORY,WEIGHT,BUYER,SALE_DATE FROM sale WHERE strftime('%Y-%m',SALE_DATE)=strftime('%Y-%m','now') ORDER BY SALE_DATE DESC", ()),
                "i": ("SELECT i.ITEMS_JSON FROM invoice_data i JOIN sale s ON s.ID=i.SALE_ID WHERE strftime('%Y-%m',s.SALE_DATE)=strftime('%Y-%m','now')", ()),
            },
            "yearly":  {
                "p": ("SELECT ID,ITEM,MATERIAL,CATEGORY,G_WEIGHT,L_WEIGHT,N_WEIGHT,SELLER,PURCHASE_DATE FROM purchase WHERE strftime('%Y',PURCHASE_DATE)=strftime('%Y','now') ORDER BY PURCHASE_DATE DESC", ()),
                "s": ("SELECT ITEM,MATERIAL,CATEGORY,WEIGHT,BUYER,SALE_DATE FROM sale WHERE strftime('%Y',SALE_DATE)=strftime('%Y','now') ORDER BY SALE_DATE DESC", ()),
                "i": ("SELECT i.ITEMS_JSON FROM invoice_data i JOIN sale s ON s.ID=i.SALE_ID WHERE strftime('%Y',s.SALE_DATE)=strftime('%Y','now')", ()),
            },
        }
        if report_type not in QUERIES:
            report_type = "daily"
        purchase_q = QUERIES[report_type]["p"]
        sale_q     = QUERIES[report_type]["s"]
        inv_q      = QUERIES[report_type]["i"]

    purchase = conn.execute(*purchase_q).fetchall()
    sales    = conn.execute(*sale_q).fetchall()
    inv_rows = conn.execute(*inv_q).fetchall()

    # Build material breakdown AND sale total in a single pass
    mat_totals    = {}
    sale_total_rs = 0.0
    for row in inv_rows:
        try:
            for itm in json.loads(row["ITEMS_JSON"]):
                mat = itm.get("material", "OTHER")
                val = itm.get("item_total", 0)
                sale_total_rs += val
                if mat not in mat_totals:
                    mat_totals[mat] = {"material": mat, "count": 0, "weight": 0.0, "revenue": 0.0}
                mat_totals[mat]["count"]   += 1
                mat_totals[mat]["weight"]  += itm.get("weight", 0)
                mat_totals[mat]["revenue"] += val
        except Exception:
            pass
    material_breakdown = sorted(mat_totals.values(), key=lambda x: x["revenue"], reverse=True)

    conn.close()

    sale_total_rs     = round(sale_total_rs, 2)
    sale_total_gst    = round(sale_total_rs * 1.03, 2)
    total_sale_weight = round(sum(s["WEIGHT"] for s in sales), 3)
    total_pur_weight  = round(sum((p["N_WEIGHT"] if "N_WEIGHT" in p.keys() else p["WEIGHT"]) for p in purchase), 3)

    return render_template("report.html", purchase=purchase, sales=sales,
        report_type=report_type, date_from=date_from, date_to=date_to,
        sale_total_rs=sale_total_rs, sale_total_gst=sale_total_gst,
        total_sale_weight=total_sale_weight, total_pur_weight=total_pur_weight,
        material_breakdown=material_breakdown)



# ---------- XLSX HELPER ----------
def make_xlsx(sheets, filename):
    """sheets = list of (sheet_name, headers, rows_of_values)"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    import io as _io

    HEADER_FILL = PatternFill("solid", fgColor="8B0000")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    LIGHT_FILL  = PatternFill("solid", fgColor="F9F9F9")
    CENTER      = Alignment(horizontal="center", vertical="center")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, headers, rows in sheets:
        ws = wb.create_sheet(title=sheet_name[:31])
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = CENTER
        ws.row_dimensions[1].height = 18

        for row_idx, row in enumerate(rows, start=2):
            ws.append(list(row))
            if row_idx % 2 == 0:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = LIGHT_FILL

        for col_idx, header in enumerate(headers, start=1):
            max_len = len(header)
            for row in rows:
                val = str(row[col_idx - 1]) if row[col_idx - 1] is not None else ""
                max_len = max(max_len, len(val))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)
        ws.freeze_panes = "A2"

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ---------- EXPORT STOCK ----------
@app.route("/export_stock")
@login_required
def export_stock():
    conn    = get_db()
    rows    = conn.execute("SELECT ID,ITEM,MATERIAL,CATEGORY,G_WEIGHT,L_WEIGHT,N_WEIGHT,PURITY,CENT,MRP_PRICE,PURCHASE_DATE,NOTES FROM stock").fetchall()
    conn.close()
    headers = ["Tag ID","Item","Material","Category","G.Weight","L.Weight","N.Weight","Purity","Cent","MRP Price","Purchase Date","Notes"]
    data    = [[r["ID"],r["ITEM"],r["MATERIAL"],r["CATEGORY"],r["G_WEIGHT"],r["L_WEIGHT"],r["N_WEIGHT"],r["PURITY"],r["CENT"],r["MRP_PRICE"],r["PURCHASE_DATE"],r["NOTES"]] for r in rows]
    return make_xlsx([("Stock", headers, data)], f"stock_report_{date.today().strftime('%Y-%m-%d')}.xlsx")


# ---------- EXPORT REPORT ----------
@app.route("/export_report/<report_type>")
@login_required
def export_report(report_type):
    conn      = get_db()
    today     = date.today().strftime("%Y-%m-%d")
    date_from = request.args.get("date_from", "")
    date_to   = request.args.get("date_to", "")
    QUERIES = {
        "daily":   {"p": ("SELECT ID,ITEM,MATERIAL,CATEGORY,G_WEIGHT,L_WEIGHT,N_WEIGHT,PURITY,CENT,MRP_PRICE,SELLER,PHONE,PURCHASE_DATE FROM purchase WHERE PURCHASE_DATE=?", (today,)),
                    "s": ("SELECT ITEM,MATERIAL,CATEGORY,WEIGHT,PURITY,MRP_PRICE,BUYER,PHONE,SALE_DATE FROM sale WHERE SALE_DATE=?", (today,))},
        "monthly": {"p": ("SELECT ID,ITEM,MATERIAL,CATEGORY,G_WEIGHT,L_WEIGHT,N_WEIGHT,PURITY,CENT,MRP_PRICE,SELLER,PHONE,PURCHASE_DATE FROM purchase WHERE strftime('%Y-%m',PURCHASE_DATE)=strftime('%Y-%m','now')", ()),
                    "s": ("SELECT ITEM,MATERIAL,CATEGORY,WEIGHT,PURITY,MRP_PRICE,BUYER,PHONE,SALE_DATE FROM sale WHERE strftime('%Y-%m',SALE_DATE)=strftime('%Y-%m','now')", ())},
        "yearly":  {"p": ("SELECT ID,ITEM,MATERIAL,CATEGORY,G_WEIGHT,L_WEIGHT,N_WEIGHT,PURITY,CENT,MRP_PRICE,SELLER,PHONE,PURCHASE_DATE FROM purchase WHERE strftime('%Y',PURCHASE_DATE)=strftime('%Y','now')", ()),
                    "s": ("SELECT ITEM,MATERIAL,CATEGORY,WEIGHT,PURITY,MRP_PRICE,BUYER,PHONE,SALE_DATE FROM sale WHERE strftime('%Y',SALE_DATE)=strftime('%Y','now')", ())},
        "custom":  {"p": ("SELECT ID,ITEM,MATERIAL,CATEGORY,G_WEIGHT,L_WEIGHT,N_WEIGHT,PURITY,CENT,MRP_PRICE,SELLER,PHONE,PURCHASE_DATE FROM purchase WHERE PURCHASE_DATE BETWEEN ? AND ?", (date_from, date_to)),
                    "s": ("SELECT ITEM,MATERIAL,CATEGORY,WEIGHT,PURITY,MRP_PRICE,BUYER,PHONE,SALE_DATE FROM sale WHERE SALE_DATE BETWEEN ? AND ?", (date_from, date_to))},
    }
    if report_type not in QUERIES:
        return "Invalid report type.", 400
    purchase = conn.execute(*QUERIES[report_type]["p"]).fetchall()
    sales    = conn.execute(*QUERIES[report_type]["s"]).fetchall()
    conn.close()

    p_headers = ["Tag ID","Item","Material","Category","G.Weight","L.Weight","N.Weight","Purity","Cent","MRP Price","Seller","Phone","Date"]
    p_data    = [[r["ID"],r["ITEM"],r["MATERIAL"],r["CATEGORY"],r["G_WEIGHT"],r["L_WEIGHT"],r["N_WEIGHT"],r.get("PURITY"),r.get("CENT"),r.get("MRP_PRICE"),r["SELLER"],r["PHONE"],r["PURCHASE_DATE"]] for r in purchase]
    s_headers = ["Item","Material","Category","N.Weight","Purity","MRP Price","Buyer","Phone","Date"]
    s_data    = [[r["ITEM"],r["MATERIAL"],r["CATEGORY"],r["WEIGHT"],r.get("PURITY"),r.get("MRP_PRICE"),r["BUYER"],r["PHONE"],r["SALE_DATE"]] for r in sales]

    return make_xlsx([("Purchases", p_headers, p_data), ("Sales", s_headers, s_data)], f"{report_type}_report_{date.today().strftime('%Y-%m-%d')}.xlsx")


# ---------- EXPORT BY DATE RANGE ----------
@app.route("/export_by_date")
@login_required
def export_by_date():
    """
    DB FORMAT — Export by Date Range
    ==================================
    Route: /export_by_date?table=<table>&date_from=YYYY-MM-DD&date_to=YYYY-MM-DD
    Allowed tables: stock, purchase, sale
    Output: Single-sheet .xlsx — same column layout as /export/<table_name>.
    """
    table     = request.args.get("table", "stock").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to   = request.args.get("date_to", "").strip()

    if not date_from or not date_to:
        return "Date range required.", 400

    conn = get_db()

    if table == "stock":
        rows    = conn.execute(
            "SELECT ID,ITEM,MATERIAL,CATEGORY,G_WEIGHT,L_WEIGHT,N_WEIGHT,PURITY,CENT,MRP_PRICE,PURCHASE_DATE,NOTES FROM stock WHERE PURCHASE_DATE BETWEEN ? AND ? ORDER BY PURCHASE_DATE DESC",
            (date_from, date_to)).fetchall()
        conn.close()
        headers = ["Tag ID","Item","Material","Category","G.Weight","L.Weight","N.Weight","Purity","Cent","MRP Price","Purchase Date","Notes"]
        data    = [[r["ID"],r["ITEM"],r["MATERIAL"],r["CATEGORY"],r["G_WEIGHT"],r["L_WEIGHT"],r["N_WEIGHT"],r["PURITY"],r["CENT"],r["MRP_PRICE"],r["PURCHASE_DATE"],r["NOTES"]] for r in rows]
        fname   = f"stock_{date_from}_to_{date_to}.xlsx"

    elif table == "purchase":
        rows    = conn.execute(
            "SELECT ID,ITEM,MATERIAL,CATEGORY,G_WEIGHT,L_WEIGHT,N_WEIGHT,PURITY,CENT,MRP_PRICE,SELLER,PHONE,PURCHASE_DATE FROM purchase WHERE PURCHASE_DATE BETWEEN ? AND ? ORDER BY PURCHASE_DATE DESC",
            (date_from, date_to)).fetchall()
        conn.close()
        headers = ["Tag ID","Item","Material","Category","G.Weight","L.Weight","N.Weight","Purity","Cent","MRP Price","Seller","Phone","Purchase Date"]
        data    = [[r["ID"],r["ITEM"],r["MATERIAL"],r["CATEGORY"],r["G_WEIGHT"],r["L_WEIGHT"],r["N_WEIGHT"],r["PURITY"],r["CENT"],r["MRP_PRICE"],r["SELLER"],r["PHONE"],r["PURCHASE_DATE"]] for r in rows]
        fname   = f"purchase_{date_from}_to_{date_to}.xlsx"

    elif table == "sale":
        rows    = conn.execute(
            "SELECT ID,ITEM,MATERIAL,CATEGORY,WEIGHT,PURITY,MRP_PRICE,BUYER,PHONE,SALE_DATE FROM sale WHERE SALE_DATE BETWEEN ? AND ? ORDER BY SALE_DATE DESC",
            (date_from, date_to)).fetchall()
        conn.close()
        headers = ["ID","Item","Material","Category","N.Weight","Purity","MRP Price","Buyer","Phone","Sale Date"]
        data    = [[r["ID"],r["ITEM"],r["MATERIAL"],r["CATEGORY"],r["WEIGHT"],r["PURITY"],r["MRP_PRICE"],r["BUYER"],r["PHONE"],r["SALE_DATE"]] for r in rows]
        fname   = f"sales_{date_from}_to_{date_to}.xlsx"

    else:
        conn.close()
        return "Invalid table.", 400

    return make_xlsx([(table.capitalize(), headers, data)], f"{table}_{date_from}_to_{date_to}.xlsx")


# ---------- EXPORT DATA PAGE ----------
@app.route("/export_data")
@login_required
def export_data():
    return render_template("export_data.html")

@app.route("/export/<table_name>")
@login_required
def export_table(table_name):
    """
    DB FORMAT — Single Table Export
    =================================
    Route: /export/<table_name>
    Allowed tables and columns exported:
    ─────────────────────────────────────────────────────────────────
    stock      : ID, ITEM, MATERIAL, CATEGORY, G_WEIGHT, L_WEIGHT,
                 N_WEIGHT, PURITY, PURCHASE_DATE, NOTES
    purchase   : ID, ITEM, MATERIAL, CATEGORY, G_WEIGHT, L_WEIGHT,
                 N_WEIGHT, PURITY, SELLER, PHONE, PURCHASE_DATE
    sale       : ID, ITEM, MATERIAL, CATEGORY, WEIGHT, PURITY,
                 BUYER, PHONE, SALE_DATE
    payments   : SALE_ID, BUYER_NAME, BUYER_PHONE, TOTAL_AMOUNT,
                 PAID_AMOUNT, DUE_AMOUNT, SALE_DATE,
                 LAST_PAYMENT_DATE, STATUS
    trade_dues : SELLER_NAME, SELLER_PHONE, MATERIAL, AGREED_WEIGHT,
                 GIVEN_WEIGHT, DUE_WEIGHT, PURCHASE_DATE,
                 LAST_UPDATE, STATUS
    ─────────────────────────────────────────────────────────────────
    Output: .xlsx file — Row 1 = headers, Row 2+ = data.
    Can be re-imported via /import_data.
    """
    allowed = {
        "stock":      ("stock",      ["ID","ITEM","MATERIAL","CATEGORY","G_WEIGHT","L_WEIGHT","N_WEIGHT","PURITY","CENT","MRP_PRICE","PURCHASE_DATE","NOTES"]),
        "purchase":   ("purchase",   ["ID","ITEM","MATERIAL","CATEGORY","G_WEIGHT","L_WEIGHT","N_WEIGHT","PURITY","CENT","MRP_PRICE","SELLER","PHONE","PURCHASE_DATE"]),
        "sale":       ("sale",       ["ID","ITEM","MATERIAL","CATEGORY","WEIGHT","PURITY","MRP_PRICE","BUYER","PHONE","SALE_DATE"]),
        "payments":   ("payments",   ["SALE_ID","BUYER_NAME","BUYER_PHONE","TOTAL_AMOUNT","PAID_AMOUNT","DUE_AMOUNT","SALE_DATE","LAST_PAYMENT_DATE","STATUS"]),
        "trade_dues": ("trade_dues", ["SELLER_NAME","SELLER_PHONE","MATERIAL","AGREED_WEIGHT","GIVEN_WEIGHT","DUE_WEIGHT","PURCHASE_DATE","LAST_UPDATE","STATUS"]),
    }
    if table_name not in allowed:
        return "Invalid table.", 400
    table, columns = allowed[table_name]
    conn = get_db()
    rows = conn.execute(f"SELECT {','.join(columns)} FROM {table}").fetchall()
    conn.close()
    data      = [[row[col] for col in columns] for row in rows]
    return make_xlsx([(table_name.capitalize(), columns, data)], f"{table_name}_export_{date.today().strftime('%Y-%m-%d')}.xlsx")


# ---------- UNIFIED SEARCH ----------
@app.route("/search", methods=["GET", "POST"])
@login_required
def search():
    stock_results = sale_results = purchase_results = []
    search_value  = ""
    search_scope  = "all"
    if request.method == "POST":
        search_value = request.form.get("value", "").strip()
        search_scope = request.form.get("scope", "all")
        if search_value:
            like = f"%{search_value.lower()}%"
            conn = get_db()
            if search_scope in ("all", "stock"):
                stock_results = conn.execute("""
                    SELECT ID,ITEM,MATERIAL,CATEGORY,G_WEIGHT,L_WEIGHT,N_WEIGHT,PURITY,PURCHASE_DATE
                    FROM stock WHERE LOWER(ID) LIKE ? OR LOWER(ITEM) LIKE ?
                    OR LOWER(MATERIAL) LIKE ? OR LOWER(CATEGORY) LIKE ?
                """, (like,)*4).fetchall()
            if search_scope in ("all", "sales"):
                sale_results = conn.execute("""
                    SELECT ID,ITEM,MATERIAL,CATEGORY,WEIGHT,BUYER,PHONE,SALE_DATE
                    FROM sale WHERE LOWER(ITEM) LIKE ? OR LOWER(BUYER) LIKE ?
                    OR LOWER(MATERIAL) LIKE ? OR CAST(ID AS TEXT) LIKE ?
                """, (like,)*4).fetchall()
            if search_scope in ("all", "purchases"):
                purchase_results = conn.execute("""
                    SELECT ID,ITEM,MATERIAL,CATEGORY,G_WEIGHT,L_WEIGHT,N_WEIGHT,PURITY,SELLER,PHONE,PURCHASE_DATE
                    FROM purchase WHERE LOWER(ID) LIKE ? OR LOWER(ITEM) LIKE ?
                    OR LOWER(SELLER) LIKE ? OR LOWER(MATERIAL) LIKE ?
                """, (like,)*4).fetchall()
            conn.close()
    return render_template("search.html", stock_results=stock_results,
        sale_results=sale_results, purchase_results=purchase_results,
        search_value=search_value, search_scope=search_scope)


# ---------- IMPORT DATA ----------
@app.route("/import_data", methods=["GET", "POST"])
@login_required
def import_data():
    """
    DB FORMAT — Import Data
    ========================
    Accepts .csv or .xlsx files.
    For .xlsx exported by this app: Row 1 = note, Row 2 = headers, Row 3+ = data.
    For plain .csv/.xlsx: Row 1 = headers, Row 2+ = data.

    Import Types & Required Columns:
    ─────────────────────────────────────────────────────────────────
    stock      : ID, ITEM, MATERIAL, CATEGORY, G_WEIGHT (or WEIGHT),
                 L_WEIGHT, N_WEIGHT, PURITY, PURCHASE_DATE
                 Optional: NOTES
    purchase   : ID, ITEM, MATERIAL, CATEGORY, G_WEIGHT (or WEIGHT),
                 L_WEIGHT, N_WEIGHT, PURITY, SELLER, PHONE,
                 PURCHASE_DATE
    sale       : ITEM, MATERIAL, CATEGORY, WEIGHT, PURITY,
                 BUYER, PHONE, SALE_DATE  (ID auto-assigned)
    payments   : SALE_ID, BUYER_NAME, BUYER_PHONE, TOTAL_AMOUNT,
                 PAID_AMOUNT, DUE_AMOUNT, SALE_DATE,
                 LAST_PAYMENT_DATE, STATUS
    trade_dues : SELLER_NAME, SELLER_PHONE, MATERIAL, AGREED_WEIGHT,
                 GIVEN_WEIGHT, DUE_WEIGHT, PURCHASE_DATE,
                 LAST_UPDATE, STATUS
    ─────────────────────────────────────────────────────────────────
    Notes:
    - MATERIAL and CATEGORY are auto-uppercased on import.
    - Duplicate IDs are silently skipped (INSERT OR IGNORE).
    - Missing PURCHASE_DATE / SALE_DATE default to today.
    """
    message = None
    errors  = []
    if request.method == "POST":
        validate_csrf()
        import_type = request.form.get("import_type")
        file        = request.files.get("file")
        if not file or file.filename == "":
            message = "No file selected."
            return render_template("import_data.html", message=message, errors=errors)
        filename = file.filename.lower()
        try:
            rows = []
            if filename.endswith(".csv"):
                rows = list(csv.DictReader(io.StringIO(file.stream.read().decode("utf-8-sig"))))
            elif filename.endswith((".xlsx", ".xls")):
                import openpyxl
                wb   = openpyxl.load_workbook(file)
                ws   = wb.active
                # Backup xlsx files have a note in row 1 and headers in row 2.
                # Detect this by checking if row 2 contains known header keywords.
                KNOWN = {"ID","ITEM","MATERIAL","BUYER","SELLER","WEIGHT","G_WEIGHT","N_WEIGHT","SALE_DATE",
                         "PURCHASE_DATE","BUYER_NAME","SELLER_NAME","TOTAL_AMOUNT",
                         "AGREED_WEIGHT","SALE_ID"}
                row2  = [str(c.value).strip().upper() if c.value else "" for c in ws[2]]
                if any(v in KNOWN for v in row2):
                    heads    = [str(c.value).strip() if c.value else "" for c in ws[2]]
                    data_rows = ws.iter_rows(min_row=3, values_only=True)
                else:
                    heads    = [str(c.value).strip() if c.value else "" for c in ws[1]]
                    data_rows = ws.iter_rows(min_row=2, values_only=True)
                for row in data_rows:
                    if any(v is not None for v in row):
                        rows.append(dict(zip(heads, row)))
            else:
                message = "Only .csv or .xlsx files are supported."
                return render_template("import_data.html", message=message, errors=errors)
        except Exception as e:
            message = f"Could not read file: {str(e)}"
            return render_template("import_data.html", message=message, errors=errors)

        conn    = get_db()
        cur     = conn.cursor()
        success = 0
        today   = datetime.now().strftime("%Y-%m-%d")

        for idx, row in enumerate(rows, start=2):
            try:
                row = {k.strip().upper(): str(v).strip() if v is not None else "" for k, v in row.items()}

                if import_type == "stock":
                    mat = row["MATERIAL"].upper()
                    purity_val = row.get("PURITY","").strip()
                    purity_imp = None if (mat == "SILVER" and purity_val == "") else float(purity_val or 0)
                    cent_val   = row.get("CENT","").strip()
                    cent_imp   = float(cent_val) if (mat == "DIAMOND" and cent_val) else None
                    mrp_val    = row.get("MRP_PRICE","").strip()
                    mrp_imp    = float(mrp_val) if (mat == "SILVER" and mrp_val) else None
                    cur.execute(
                        "INSERT OR IGNORE INTO stock (ID,ITEM,MATERIAL,CATEGORY,G_WEIGHT,L_WEIGHT,N_WEIGHT,PURITY,PURCHASE_DATE,NOTES,CENT,MRP_PRICE) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                        (row["ID"], row["ITEM"], mat, row["CATEGORY"].upper(),
                         float(row.get("G_WEIGHT") or row.get("G.WEIGHT") or row.get("WEIGHT",0)),
                         float(row.get("L_WEIGHT") or row.get("L.WEIGHT") or 0),
                         float(row.get("N_WEIGHT") or row.get("N.WEIGHT") or row.get("WEIGHT",0)),
                         purity_imp,
                         row.get("PURCHASE_DATE") or today,
                         row.get("NOTES", ""), cent_imp, mrp_imp))

                elif import_type == "purchase":
                    mat = row["MATERIAL"].upper()
                    purity_val = row.get("PURITY","").strip()
                    purity_imp = None if (mat == "SILVER" and purity_val == "") else float(purity_val or 0)
                    cent_val   = row.get("CENT","").strip()
                    cent_imp   = float(cent_val) if (mat == "DIAMOND" and cent_val) else None
                    mrp_val    = row.get("MRP_PRICE","").strip()
                    mrp_imp    = float(mrp_val) if (mat == "SILVER" and mrp_val) else None
                    cur.execute(
                        "INSERT OR IGNORE INTO purchase (ID,ITEM,MATERIAL,CATEGORY,G_WEIGHT,L_WEIGHT,N_WEIGHT,PURITY,SELLER,PHONE,PURCHASE_DATE,CENT,MRP_PRICE) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                        (row["ID"], row["ITEM"], mat, row["CATEGORY"].upper(),
                         float(row.get("G_WEIGHT") or row.get("G.WEIGHT") or row.get("WEIGHT",0)),
                         float(row.get("L_WEIGHT") or row.get("L.WEIGHT") or 0),
                         float(row.get("N_WEIGHT") or row.get("N.WEIGHT") or row.get("WEIGHT",0)),
                         purity_imp, row["SELLER"], row["PHONE"],
                         row.get("PURCHASE_DATE") or today, cent_imp, mrp_imp))

                elif import_type == "sale":
                    mat = row["MATERIAL"].upper()
                    purity_val = row.get("PURITY","").strip()
                    purity_imp = None if (mat == "SILVER" and purity_val == "") else float(purity_val or 0)
                    mrp_val    = row.get("MRP_PRICE","").strip()
                    mrp_imp    = float(mrp_val) if (mat == "SILVER" and mrp_val) else None
                    cur.execute(
                        "INSERT INTO sale (ITEM,MATERIAL,CATEGORY,WEIGHT,PURITY,BUYER,PHONE,SALE_DATE,BUYER_ADDRESS,MRP_PRICE) VALUES (?,?,?,?,?,?,?,?,'',?)",
                        (row["ITEM"], mat, row["CATEGORY"].upper(),
                         float(row["WEIGHT"]), purity_imp,
                         row["BUYER"], row["PHONE"],
                         row.get("SALE_DATE") or today, mrp_imp))

                elif import_type == "payments":
                    # Skip already-cleared if STATUS column says Cleared and DUE is 0
                    status = row.get("STATUS", "Pending")
                    cur.execute(
                        "INSERT OR IGNORE INTO payments (SALE_ID,BUYER_NAME,BUYER_PHONE,TOTAL_AMOUNT,PAID_AMOUNT,DUE_AMOUNT,SALE_DATE,LAST_PAYMENT_DATE,STATUS) VALUES (?,?,?,?,?,?,?,?,?)",
                        (int(float(row["SALE_ID"])), row["BUYER_NAME"], row["BUYER_PHONE"],
                         float(row["TOTAL_AMOUNT"]), float(row["PAID_AMOUNT"]),
                         float(row["DUE_AMOUNT"]),
                         row.get("SALE_DATE") or today,
                         row.get("LAST_PAYMENT_DATE") or today,
                         status))

                elif import_type == "trade_dues":
                    status = row.get("STATUS", "Pending")
                    cur.execute(
                        "INSERT INTO trade_dues (SELLER_NAME,SELLER_PHONE,MATERIAL,AGREED_WEIGHT,GIVEN_WEIGHT,DUE_WEIGHT,PURCHASE_DATE,LAST_UPDATE,STATUS) VALUES (?,?,?,?,?,?,?,?,?)",
                        (row["SELLER_NAME"], row["SELLER_PHONE"], row["MATERIAL"].upper(),
                         float(row["AGREED_WEIGHT"]), float(row["GIVEN_WEIGHT"]),
                         float(row["DUE_WEIGHT"]),
                         row.get("PURCHASE_DATE") or today,
                         row.get("LAST_UPDATE") or today,
                         status))

                success += 1
            except Exception as e:
                errors.append(f"Row {idx}: {str(e)}")

        conn.commit()
        conn.close()
        audit("IMPORT", f"{import_type} — {success} records imported")
        message = f"Successfully imported {success} records."
        if errors:
            message += f" {len(errors)} rows had errors and were skipped."
    return render_template("import_data.html", message=message, errors=errors)

# ---------- DUE PAYMENTS ----------
@app.route("/due_payments")
@login_required
def due_payments():
    PER_PAGE = 50
    try:
        page = max(1, int(request.args.get("page", 1)))
    except ValueError:
        page = 1
    q             = request.args.get("q", "").strip()
    status_filter = request.args.get("status", "").strip()
    conn   = get_db()
    params = []
    where  = []
    if q:
        like = f"%{q.lower()}%"
        where.append("(LOWER(BUYER_NAME) LIKE ? OR BUYER_PHONE LIKE ?)")
        params.extend([like, like])
    if status_filter:
        where.append("STATUS = ?")
        params.append(status_filter)
    where_clause = ("WHERE " + " AND ".join(where)) if where else ""
    total       = conn.execute(f"SELECT COUNT(*) FROM payments {where_clause}", params).fetchone()[0]
    total_pages = max(1, (total + PER_PAGE - 1) // PER_PAGE)
    page        = min(page, total_pages)
    payments_rows = conn.execute(
        f"SELECT * FROM payments {where_clause} ORDER BY STATUS ASC, SALE_DATE DESC LIMIT ? OFFSET ?",
        params + [PER_PAGE, (page-1)*PER_PAGE]).fetchall()
    # Summary stats (always from full pending set)
    summary = conn.execute("SELECT COUNT(*), IFNULL(SUM(DUE_AMOUNT),0) FROM payments WHERE STATUS='Pending'").fetchone()
    pending_count   = summary[0]
    total_pending   = round(summary[1], 2)
    conn.close()
    return render_template("due_payments.html", payments=payments_rows,
        page=page, total_pages=total_pages, total=total,
        q=q, status_filter=status_filter,
        pending_count=pending_count, total_pending=total_pending)

@app.route("/add_payment/<int:payment_id>", methods=["POST"])
@login_required
def add_payment(payment_id):
    validate_csrf()
    conn = get_db()
    row  = conn.execute("SELECT * FROM payments WHERE ID=?", (payment_id,)).fetchone()
    if not row:
        conn.close()
        return redirect("/due_payments")
    try:
        amount = float(request.form.get("amount", 0))
    except ValueError:
        conn.close()
        return redirect("/due_payments")
    if amount <= 0:
        conn.close()
        return redirect("/due_payments")
    new_paid = round(row["PAID_AMOUNT"] + amount, 2)
    new_due  = round(row["TOTAL_AMOUNT"] - new_paid, 2)
    if new_due < 0:
        new_due  = 0
        new_paid = row["TOTAL_AMOUNT"]
    conn.execute("UPDATE payments SET PAID_AMOUNT=?,DUE_AMOUNT=?,LAST_PAYMENT_DATE=?,STATUS=? WHERE ID=?",
        (new_paid, new_due, datetime.now().strftime("%Y-%m-%d"),
         "Cleared" if new_due <= 0 else "Pending", payment_id))
    conn.commit()
    conn.close()
    return redirect("/due_payments")


# ---------- REPRINT INVOICE ----------
@app.route("/reprint_invoice/<int:sale_id>")
@login_required
def reprint_invoice(sale_id):
    conn = get_db()
    row  = conn.execute("SELECT * FROM invoice_data WHERE SALE_ID=?", (sale_id,)).fetchone()
    conn.close()
    if not row:
        return "Invoice data not found.", 404
    invoice_path = None
    try:
        raw_items = json.loads(row["ITEMS_JSON"])
        # Normalise — old saved carts may not have flat_price key
        for itm in raw_items:
            itm.setdefault("flat_price", 0)
            itm.setdefault("rate_per_gram", 0)
            itm.setdefault("making_charges", 0)
            itm.setdefault("cent", None)
            itm.setdefault("mrp_price", None)
        invoice_path = generate_invoice(sale_id=sale_id, buyer_name=row["BUYER_NAME"],
            buyer_phone=row["BUYER_PHONE"], buyer_address=row["BUYER_STATE"],
            buyer_gstin=row["BUYER_GSTIN"], payment_method=row["PAYMENT_METHOD"],
            sale_date=row["SALE_DATE"], items=raw_items)
        with open(invoice_path, "rb") as f:
            pdf_data = f.read()
        return Response(pdf_data, mimetype="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={os.path.basename(invoice_path)}"})
    except Exception as e:
        return f"Could not generate invoice: {str(e)}"
    finally:
        if invoice_path and os.path.exists(invoice_path):
            try:
                os.remove(invoice_path)
            except OSError:
                pass


# ---------- EDIT SALE (buyer info + date only) ----------
@app.route("/edit_sale/<int:sale_id>", methods=["GET", "POST"])
@login_required
def edit_sale(sale_id):
    conn = get_db()
    sale = conn.execute("SELECT * FROM sale WHERE ID=?", (sale_id,)).fetchone()
    if not sale:
        conn.close()
        return redirect("/view_sales")
    error = None
    if request.method == "POST":
        validate_csrf()
        buyer         = request.form.get("buyer", "").strip()[:100]
        phone         = clean_phone(request.form.get("phone", "").strip())
        buyer_address = request.form.get("buyer_address", "").strip()[:200]
        sale_date     = request.form.get("sale_date", "").strip()
        if not buyer:
            error = "Buyer name is required."
        elif not phone.isdigit() or len(phone) != 10:
            error = "Phone must be exactly 10 digits."
        elif not sale_date:
            error = "Sale date is required."
        else:
            try:
                conn.execute(
                    "UPDATE sale SET BUYER=?,PHONE=?,BUYER_ADDRESS=?,SALE_DATE=? WHERE ID=?",
                    (buyer, phone, buyer_address, sale_date, sale_id))
                conn.execute(
                    "UPDATE payments SET BUYER_NAME=?,BUYER_PHONE=? WHERE SALE_ID=?",
                    (buyer, phone, sale_id))
                conn.execute(
                    "UPDATE invoice_data SET BUYER_NAME=?,BUYER_PHONE=?,BUYER_STATE=? WHERE SALE_ID=?",
                    (buyer, phone, buyer_address, sale_id))
                conn.commit()
                audit("EDIT_SALE", f"Sale #{sale_id} buyer/date updated")
                conn.close()
                return redirect("/view_sales")
            except Exception as e:
                conn.rollback()
                error = f"Update failed: {str(e)}"
    conn.close()
    return render_template("edit_sale.html", sale=sale, error=error)


# ---------- DELETE SALE (restores stock) ----------
@app.route("/delete_sale/<int:sale_id>", methods=["POST"])
@login_required
def delete_sale(sale_id):
    validate_csrf()
    conn = get_db()
    cur  = conn.cursor()
    try:
        inv_row = cur.execute("SELECT * FROM invoice_data WHERE SALE_ID=?", (sale_id,)).fetchone()
        if inv_row:
            for itm in json.loads(inv_row["ITEMS_JSON"]):
                if not cur.execute("SELECT ID FROM stock WHERE ID=?", (itm["stock_id"],)).fetchone():
                    pur = cur.execute("SELECT PURCHASE_DATE FROM purchase WHERE ID=?", (itm["stock_id"],)).fetchone()
                    cur.execute("INSERT INTO stock (ID,ITEM,MATERIAL,CATEGORY,G_WEIGHT,L_WEIGHT,N_WEIGHT,PURITY,PURCHASE_DATE,NOTES) VALUES (?,?,?,?,?,?,?,?,?,?)",
                        (itm["stock_id"], itm["item"], itm["material"], itm["category"],
                         itm.get("g_weight", itm.get("weight",0)), itm.get("l_weight",0), itm.get("weight",0), itm["purity"],
                         pur["PURCHASE_DATE"] if pur else date.today().strftime("%Y-%m-%d"), ""))
        cur.execute("DELETE FROM sale WHERE ID=?", (sale_id,))
        cur.execute("DELETE FROM invoice_data WHERE SALE_ID=?", (sale_id,))
        cur.execute("DELETE FROM payments WHERE SALE_ID=?", (sale_id,))
        conn.commit()
        audit("DELETE_SALE", f"Sale #{sale_id} deleted — stock restored")
    except Exception:
        conn.rollback()
    finally:
        conn.close()
    return redirect("/view_sales")


# ---------- DELETE PURCHASE (blocked if item was sold) ----------
@app.route("/delete_purchase/<path:purchase_id>", methods=["POST"])
@login_required
def delete_purchase(purchase_id):
    validate_csrf()
    conn = get_db()
    sold = conn.execute(
        "SELECT ID FROM sale s JOIN invoice_data i ON s.ID=i.SALE_ID WHERE i.ITEMS_JSON LIKE ?",
        (f'%"{purchase_id}"%',)).fetchone()
    if sold:
        conn.close()
        return """<html><body style='font-family:sans-serif;padding:40px'>
            <h3 style='color:red'>⚠️ Cannot Delete Purchase</h3>
            <p>This item has already been sold. Deleting its purchase record would corrupt your history.</p>
            <a href='/view_purchases' style='background:#8B0000;color:white;padding:8px 16px;border-radius:4px;text-decoration:none'>← Go Back</a>
        </body></html>""", 400
    conn.execute("DELETE FROM purchase WHERE ID=?", (purchase_id,))
    conn.commit()
    conn.close()
    return redirect("/view_purchases")


# ---------- DELETE STOCK ----------
@app.route("/delete_stock/<path:stock_id>", methods=["POST"])
@login_required
def delete_stock(stock_id):
    validate_csrf()
    conn = get_db()
    conn.execute("DELETE FROM stock WHERE ID=?", (stock_id,))
    conn.commit()
    conn.close()
    audit("DELETE_STOCK", f"Stock item {stock_id} deleted")
    return redirect("/view_stock")


# ---------- EDIT STOCK ----------
@app.route("/edit_stock/<path:stock_id>", methods=["GET", "POST"])
@login_required
def edit_stock(stock_id):
    conn = get_db()
    item = conn.execute("SELECT * FROM stock WHERE ID=?", (stock_id,)).fetchone()
    if not item:
        conn.close()
        return redirect("/view_stock")
    error = None
    if request.method == "POST":
        validate_csrf()
        try:
            new_item     = request.form["item"].strip()
            new_material = request.form["material"].strip().upper()
            new_category = request.form["category"].strip().upper()
            new_g_weight = round(float(request.form["g_weight"]), 3)
            new_l_weight = round(float(request.form.get("l_weight", 0)), 3)
            new_n_weight = round(new_g_weight - new_l_weight, 3)
            # Purity: optional (None) for SILVER
            purity_raw   = request.form.get("purity", "").strip()
            new_purity   = None if (new_material == "SILVER" and purity_raw == "") else float(purity_raw or 0)
            new_notes    = request.form.get("notes", "").strip()[:300]
            # Cent: only for DIAMOND
            cent_raw     = request.form.get("cent", "").strip()
            new_cent     = float(cent_raw) if (new_material == "DIAMOND" and cent_raw) else None
            # MRP: only for SILVER
            mrp_raw      = request.form.get("mrp_price", "").strip()
            new_mrp      = float(mrp_raw) if (new_material == "SILVER" and mrp_raw) else None
        except (ValueError, KeyError):
            error = "Invalid input."
            return render_template("edit_stock.html", item=item, error=error)
        if new_g_weight <= 0 or new_n_weight <= 0 or not new_item or not new_material or not new_category:
            error = "All fields required. G.Weight must be positive and N.Weight must be greater than zero."
            return render_template("edit_stock.html", item=item, error=error)
        if new_material != "SILVER" and (new_purity is None or new_purity <= 0):
            error = "Purity must be positive for Gold and Diamond items."
            return render_template("edit_stock.html", item=item, error=error)
        try:
            conn.execute("UPDATE stock SET ITEM=?,MATERIAL=?,CATEGORY=?,G_WEIGHT=?,L_WEIGHT=?,N_WEIGHT=?,PURITY=?,NOTES=?,CENT=?,MRP_PRICE=? WHERE ID=?",
                (new_item, new_material, new_category, new_g_weight, new_l_weight, new_n_weight, new_purity, new_notes, new_cent, new_mrp, stock_id))
            conn.execute("UPDATE purchase SET ITEM=?,MATERIAL=?,CATEGORY=?,G_WEIGHT=?,L_WEIGHT=?,N_WEIGHT=?,PURITY=?,CENT=?,MRP_PRICE=? WHERE ID=?",
                (new_item, new_material, new_category, new_g_weight, new_l_weight, new_n_weight, new_purity, new_cent, new_mrp, stock_id))
            conn.commit()
            audit("EDIT_STOCK", f"Stock {stock_id} updated")
        except Exception as e:
            conn.rollback()
            error = f"Update failed: {str(e)}"
            return render_template("edit_stock.html", item=item, error=error)
        finally:
            conn.close()
        return redirect("/view_stock")
    conn.close()
    return render_template("edit_stock.html", item=item, error=error)


# ---------- AUDIT LOG VIEW ----------
@app.route("/audit_log")
@admin_required
def audit_log():
    PER_PAGE = 100
    try:
        page = max(1, int(request.args.get("page", 1)))
    except ValueError:
        page = 1
    q      = request.args.get("q", "").strip()
    conn   = get_db()
    params = []
    where  = []
    if q:
        like = f"%{q.lower()}%"
        where.append("(LOWER(USERNAME) LIKE ? OR LOWER(ACTION) LIKE ? OR LOWER(DETAIL) LIKE ?)")
        params.extend([like, like, like])
    where_clause = ("WHERE " + " AND ".join(where)) if where else ""
    total       = conn.execute(f"SELECT COUNT(*) FROM audit_log {where_clause}", params).fetchone()[0]
    total_pages = max(1, (total + PER_PAGE - 1) // PER_PAGE)
    page        = min(page, total_pages)
    logs = conn.execute(
        f"SELECT * FROM audit_log {where_clause} ORDER BY TIMESTAMP DESC LIMIT ? OFFSET ?",
        params + [PER_PAGE, (page-1)*PER_PAGE]).fetchall()
    conn.close()
    return render_template("audit_log.html", logs=logs,
        page=page, total_pages=total_pages, total=total, q=q)


# ---------- ERROR PAGES ----------
@app.errorhandler(404)
def not_found(e):
    return render_template("404.html"), 404

@app.errorhandler(403)
def forbidden(e):
    return render_template("403.html"), 403

@app.errorhandler(500)
def server_error(e):
    return render_template("500.html"), 500


# ---------- MAIN ----------
init_db()

if __name__ == "__main__":
    app.run(debug=False)